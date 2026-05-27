from io import BytesIO
from datetime import date
from fastapi import APIRouter, Depends, Query, UploadFile, File, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
import openpyxl
from ..database import get_db
from ..models.product import Product
from ..models.order import Order
from ..models.client import Client
from ..models.batch import Batch
from .deps import current_user
from ..models.user import User

router = APIRouter(tags=["export"])


def _excel_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/export/products")
async def export_products(db: AsyncSession = Depends(get_db), _: User = Depends(current_user)):
    result = await db.execute(
        select(Product).options(selectinload(Product.batches), selectinload(Product.category), selectinload(Product.supplier))
        .where(Product.is_active == True)
    )
    products = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["ID", "Barcode", "Name", "Category", "Supplier", "Cost Price ($)", "Selling Price ($)", "Min Stock", "Total Stock"])
    for p in products:
        stock = sum(b.quantity for b in p.batches)
        ws.append([
            p.id, p.barcode, p.name,
            p.category.name if p.category else "",
            p.supplier.name if p.supplier else "",
            p.cost_price / 100,
            p.selling_price / 100,
            p.min_stock_threshold,
            stock,
        ])
    return _excel_response(wb, "products.xlsx")


@router.get("/export/orders")
async def export_orders(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(current_user),
):
    q = select(Order).options(selectinload(Order.client)).order_by(Order.order_date.desc())
    result = await db.execute(q)
    orders = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["Order ID", "Client", "Date", "Status", "Total ($)", "Paid ($)", "Balance Due ($)"])
    for o in orders:
        ws.append([
            o.id,
            o.client.name if o.client else "",
            o.order_date.date().isoformat(),
            o.status,
            o.total_amount / 100,
            o.paid_amount / 100,
            o.balance_due / 100,
        ])
    return _excel_response(wb, "orders.xlsx")


@router.get("/export/clients")
async def export_clients(db: AsyncSession = Depends(get_db), _: User = Depends(current_user)):
    result = await db.execute(select(Client).where(Client.is_active == True))
    clients = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    ws.append(["ID", "Name", "Type", "Phone", "Email", "Address", "Outstanding Balance ($)"])
    for c in clients:
        ws.append([c.id, c.name, c.type, c.phone or "", c.email or "", c.address or "", c.outstanding_balance / 100])
    return _excel_response(wb, "clients.xlsx")


@router.post("/import/products", status_code=status.HTTP_200_OK)
async def import_products(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(current_user),
):
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    created = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            barcode, name, cost_price_dollars, selling_price_dollars, min_stock = row[0], row[1], row[2], row[3], row[4]
            if not name:
                continue
            product = Product(
                barcode=str(barcode) if barcode else None,
                name=str(name),
                cost_price=int(float(cost_price_dollars or 0) * 100),
                selling_price=int(float(selling_price_dollars or 0) * 100),
                min_stock_threshold=int(min_stock or 0),
            )
            db.add(product)
            created += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})
    await db.commit()
    return {"created": created, "errors": errors}


@router.post("/orders/{order_id}/invoice")
async def generate_invoice(order_id: int, _: User = Depends(current_user)):
    raise HTTPException(status_code=status.HTTP_501_NOT_IMPLEMENTED, detail="Invoice PDF generation not yet implemented")
